import os
import openpyxl

'''
根据表格里面的片段信息和剧集信息，每个dialog保存为一个文件。
切割的命令: 
时间格式: hh:mm:ss:xxx
'''

def calc_time(time):
    str_time = str(time)
    hour, minute, second, ms = str_time.split(':')
    return float(minute) * 60 + float(second) + float(ms) * 0.01

def analyse_xls(filepath, skip_rows=1):
    '''
    统计包含多少部电视剧，多少个对话，总的时间长度。
    对于情感分布该如何统计？
    :param filepath:
    :param sheetname: sheetname
    :return: list types of all instances
    '''
    workbook = openpyxl.load_workbook(filepath)
    all_sheetnames = workbook.sheetnames
    sheetname_count = 0
    dialog_count = 0
    dur_count = 0
    emo2dialog = {}
    print(all_sheetnames)
    for sheetname in all_sheetnames:
        if sheetname == 'siteng' or sheetname == '推荐电视剧列表':
            continue
        print('current sheetname {}'.format(sheetname))
        sheetname_count += 1
        booksheet = workbook[sheetname]
        rows = booksheet.rows
        all_rows = [r for r in rows]
        all_instances = all_rows[skip_rows:]
        emocolumn_instance = all_rows[0][4]
        emoindex = emocolumn_instance.value
        print('EmoIndex: {}'.format(emoindex))
        for instance in all_instances:
            Index, Episode,	startTime, endTime, EmoInfo = instance[:5]
            dialog_count += 1
            index = Index.value
            episode = Episode.value
            # 空行处理
            if index is None or episode is None:
                break
            if emoindex is not None:
                emoinfo = EmoInfo.value
                if '+' in emoinfo:
                    emos = emoinfo.replace(' ', '').split('+')
                else:
                    emos = [emoinfo]
                for emo in emos:
                    if emo2dialog.get(emo) is None:
                        emo2dialog[emo] = 1
                    else:
                        emo2dialog[emo] += 1
            start_time = startTime.value
            end_time = endTime.value
            # print(index, episode)
            # print(start_time, end_time)
            duration = calc_time(end_time) - calc_time(start_time)
            dur_count += duration
    print('total sheetname {}'.format(sheetname_count))
    print('total dialogs {}'.format(dialog_count))
    print('total duration {}'.format(dur_count))
    print(emo2dialog)
raw_movies_dir = '/Users/jinming/Desktop/works/memoconv_rawmovies'
conv_movies_dir = '/Users/jinming/Desktop/works/memoconv_convs'
segment_info_path = os.path.join(raw_movies_dir, '多模态对话数据集对话选取.xlsx')
analyse_xls(segment_info_path)