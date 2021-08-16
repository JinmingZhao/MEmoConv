import os
import math
import numpy as np
import collections
import openpyxl
import sklearn.metrics
from FileOps import read_xls, write_xls, read_file
from collections import Counter


'''
step1: 验证三个文件的行数是否一致
step2: 验证三个文件的对话的个数以及每个对话的句子数目，以及句子的ID是否唯一
step3: 验证说话人个数是否是只有AB, 验证第一句话是否为A，统计对话的轮数
step4: 然后验证情感的标注信息，所有的句子数目，完全相同的情感句子数目，第一情感相同的句子数据，情感有交集的句子数目.
step5: 文本数据的清洗，将所有的英文标点符合替换为中文标点符号, 以anno3为标准。另外整理三个人的标注信息。
step6: 根据什么策略决定最终的标注。 参考一下多情感标注的论文。
step7. 整理整个数据集的情况，并进行整理 statistic.xlsx
'''

def check_dialog(anno_instances):
    dialogs = collections.OrderedDict()
    dialogs2spks = collections.OrderedDict()
    for instance_id in range(len(anno_instances)):
        instance = anno_instances[instance_id]
        UtteranceId = instance[0].value
        if UtteranceId is None:
            continue
        dialog_id = '_'.join(UtteranceId.split('_')[:2])
        spkId = instance[4].value.replace(' ', '')
        if spkId not in ['A', 'B']:
            print('\t Error dialog {} has more than 2 speakers'.format(dialog_id))      
            exit(0)
        if dialogs.get(dialog_id) is None:
            dialogs[dialog_id] = [UtteranceId]
            if spkId != 'A':
                print('\t Error dialog {} not start at speaker A'.format(dialog_id))   
                exit(0)
            dialogs2spks[dialog_id] = [spkId]
        else:
            dialogs[dialog_id] += [UtteranceId]
            dialogs2spks[dialog_id] += [spkId]
    num_dialogs = len(dialogs)
    # print('\t dialog nums {}'.format(num_dialogs))
    num_utts = 0
    for dialog_id in dialogs.keys():
        utts = dialogs[dialog_id]
        if len(set(utts)) != len(utts):
            print('In dialog {} the uttId is not unique {} {}'.format(dialog_id, len(set(utts)),  len(utts)))
            exit(0)
        num_utts += len(utts)
    # print('\t total utts nums {}'.format(num_utts))
    total_num_turns = 0
    dialogid2num_turns = collections.OrderedDict()
    for dialog_id in dialogs2spks.keys():
        spks = dialogs2spks[dialog_id]
        num_turns = 0
        cur_spk = 'A'
        for i in range(len(spks)):
            if spks[i] != cur_spk:
                num_turns +=1 
                cur_spk = spks[i]
        dialogid2num_turns[dialog_id] = num_turns
        total_num_turns += num_turns
    # print('\t total turns nums {}'.format(num_turns))
    return num_dialogs, total_num_turns, num_utts, dialogid2num_turns, dialogs

def clean_text(utterance):
    # 将英文的标点符号修改为中文的标点符号，包括 () , . ? ! 等
    # 将句子中的空格，替换为逗号
    utterance = utterance.replace(',', '，').replace('.', '。').replace('?', '？').replace('!', '！').replace('(', '（').replace(')', '）')
    utterance = utterance.replace(' ', '，')
    return utterance

def correct_emos(emos):
    emo_list = ['Happy', 'Neutral', 'Sad', 'Anger', 'Disgust', 'Fear', 'Surprise', 'Other']
    emo_map = {'Angry': 'Anger', 'Anger ': 'Anger', 'Surprised': 'Surprise', 'other': 'Other', 'Othe': 'Other', 
                'surprise': 'Surprise', 'Surperised': 'Surprise', 'Happpy': 'Happy'}
    new_emos = []
    for emo in emos:
        if emo not in emo_list:
            if emo_map.get(emo) is not None:
                new_emos.append(emo_map[emo])
            else:
                print('Error emo name {}'.format(emo))
                exit(0)
        else:
            new_emos.append(emo)
    return new_emos

def analyse_dialog_emotions(anno_instances):
    # 注意存在情感类别拼写错误的问题，比如 Anger 和 Angry Surprised 和 Surprise
    dialogs2emos = collections.OrderedDict()
    pre_emos = None
    total_emo_anno = 0
    for instance_id in range(len(anno_instances)):
        instance = anno_instances[instance_id]
        UtteranceId = instance[0].value
        if UtteranceId is None:
            continue
        emos_str = instance[5].value
        dialog_id = '_'.join(UtteranceId.split('_')[:2])
        if dialogs2emos.get(dialog_id) is None:
            # 每个对话的第一句话的情感标注不可以为空
            if emos_str is None:
                print('\t \t Error emo anno is Empty in {}'.format(dialog_id))
                exit(0)
            emos = emos_str.split(',')
            emos = correct_emos(emos)
            dialogs2emos[dialog_id] = [emos]
            total_emo_anno += len(emos)
            # 更新 pre_emos
            pre_emos = emos
        else:
            if emos_str is None:
                emos = pre_emos
            else:
                # 随时更新 pre_emos
                emos = emos_str.split(',')
                emos = correct_emos(emos)
                pre_emos = emos
            dialogs2emos[dialog_id] += [emos]
            total_emo_anno += len(emos)
    return total_emo_anno, dialogs2emos

def analyse_emotion_three_consistency(anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos):
    '''
    三个人:
        完全相同的情感句子数目: num_total_same
        第一情感相同的句子数据: num_main_same
        情感有交集的句子数目: num_part_same
    '''
    num_total_same = 0
    num_main_same = 0
    num_part_same = 0
    for dialog_id in anno1_dialogs2emos.keys():
        anno1_dailog_emos = anno1_dialogs2emos[dialog_id]
        anno2_dailog_emos = anno2_dialogs2emos[dialog_id]
        anno3_dailog_emos = anno3_dialogs2emos[dialog_id]
        for utt_index in range(len(anno1_dailog_emos)):
            if anno1_dailog_emos[utt_index] == anno2_dailog_emos[utt_index] == anno3_dailog_emos[utt_index]:
                num_total_same += 1
            elif anno1_dailog_emos[utt_index][0] == anno2_dailog_emos[utt_index][0] == anno3_dailog_emos[utt_index][0]:
                num_main_same += 1
            else:
                part_same_emos = set(anno1_dailog_emos[utt_index]).intersection(set(anno2_dailog_emos[utt_index]),set(anno3_dailog_emos[utt_index]))
                if len(part_same_emos) > 0:
                    num_part_same += 1
    num_part_same = num_total_same + num_main_same + num_part_same
    num_main_same = num_total_same + num_main_same
    return num_total_same, num_main_same, num_part_same

def analyse_emotion_two_consistency(anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos):
    '''
    一般来说我们情感标注，只要两人的情感一致就可以进行决策。
    两个人的情感满足相似即可:
        完全相同的情感句子数目: num_total_same
        第一情感相同的句子数据: num_main_same
        情感有交集的句子数目: num_part_same
    '''
    num_total_same = 0
    num_main_same = 0
    num_part_same = 0
    for dialog_id in anno1_dialogs2emos.keys():
        anno1_dailog_emos = anno1_dialogs2emos[dialog_id]
        anno2_dailog_emos = anno2_dialogs2emos[dialog_id]
        anno3_dailog_emos = anno3_dialogs2emos[dialog_id]
        for utt_index in range(len(anno1_dailog_emos)):
            # 统计二者完全相同的
            if anno1_dailog_emos[utt_index] == anno2_dailog_emos[utt_index]:
                num_total_same += 1
            elif anno1_dailog_emos[utt_index] == anno3_dailog_emos[utt_index]:
                num_total_same += 1
            elif anno2_dailog_emos[utt_index] == anno3_dailog_emos[utt_index]:
                num_total_same += 1
            else:
                pass
            # 统计二者主情感相同
            if anno1_dailog_emos[utt_index][0] == anno2_dailog_emos[utt_index][0]:
                num_main_same += 1
            elif anno1_dailog_emos[utt_index][0] == anno3_dailog_emos[utt_index][0]:
                num_main_same += 1
            elif anno2_dailog_emos[utt_index][0] == anno3_dailog_emos[utt_index][0]:
                num_main_same += 1
            else:
                pass
            # 统计二者部分情感相同
            inter_anno12 = set(anno1_dailog_emos[utt_index]).intersection(set(anno2_dailog_emos[utt_index]))
            inter_anno13 = set(anno1_dailog_emos[utt_index]).intersection(set(anno3_dailog_emos[utt_index]))
            inter_anno23 = set(anno2_dailog_emos[utt_index]).intersection(set(anno3_dailog_emos[utt_index]))
            if len(inter_anno12) > 0 or len(inter_anno13) > 0 or  len(inter_anno23) > 0:
                num_part_same += 1
    return num_total_same, num_main_same, num_part_same



def get_movie_statistic(movie_name, anno3_num_dialogs, anno3_num_turns, anno3_num_utts, 
                                    p3_num_total_same, p3_num_main_same, p3_num_part_same,
                                    p2_num_total_same, p2_num_main_same, p2_num_part_same, 
                                    fleiss_kappa, not_sure_count):
    instance = [movie_name, anno3_num_dialogs, anno3_num_turns, anno3_num_utts, round(anno3_num_turns/anno3_num_dialogs, 2), round(anno3_num_utts/anno3_num_turns, 2)]
    instance.extend([not_sure_count, round(fleiss_kappa, 2)])
    instance.extend([p3_num_total_same, round(p3_num_total_same/anno3_num_utts, 2)])
    instance.extend([p3_num_main_same, round(p3_num_main_same/anno3_num_utts, 2)])
    instance.extend([p3_num_part_same, round(p3_num_part_same/anno3_num_utts, 2)])
    instance.extend([p2_num_total_same, round(p2_num_total_same/anno3_num_utts, 2)])
    instance.extend([p2_num_main_same, round(p2_num_main_same/anno3_num_utts, 2)])
    instance.extend([p2_num_part_same, round(p2_num_part_same/anno3_num_utts, 2)])
    return instance

def write_xls_oneline(filepath, instance):
    '''
    电视剧名字作为ID，统计每部电视剧的统计信息，如果存在那么更新，如果不存在那么追加
    head = ['movieName', 'num_dialog', 'num_turns', 'num_utts', 'numturn_per_dialog', 'numutt_per_turn', not_sure_count, fleiss_kappa,
    'total3_consist', 'total3_consist_rate', 'main3_consist', 'main3_consist_rate', 'part3_consist', 'part3_consist_rate', 'total2_consist', 'total2_consist_rate', 'main2_consist', 'main2_consist_rate', 'part2_consist', 'part2_consist_rate']
    booksheet.append(head)
    '''
    movie_name = instance[0]
    workbook = openpyxl.load_workbook(filepath)
    booksheet = workbook['statistic']
    rows = booksheet.rows
    rows = [r for r in rows]
    update_flag = False
    for i in range(len(rows)):
        if movie_name == rows[i][0].value:
            assert len(rows[i]) == len(instance)
            update_flag = True
            # start from 1
            update_row_idx = i+1
            booksheet.delete_rows(idx=update_row_idx, amount=1)
    # head = ['movieName', 'num_dialog', 'num_turns', 'num_utts', 'numutt_per_turn', 'fleiss_kappa', 'numturn_per_dialog', 'total3_consist', 'total3_consist_rate', 'main3_consist', 'main3_consist_rate', 'part3_consist', 'part3_consist_rate', 'total2_consist', 'total2_consist_rate', 'main2_consist', 'main2_consist_rate', 'part2_consist', 'part2_consist_rate']
    # booksheet.append(head)
    if update_flag is False:
        booksheet.append(instance)
        print('add the movie {}'.format(movie_name))
    else:
        print('update the movie {}'.format(movie_name))
        for colunm in range(1, len(instance)+1):
            # start from 1
            booksheet.cell(row=update_row_idx, column=colunm).value=instance[colunm-1]
    # save
    workbook.save(filepath)

def get_clean_text_dialog(save_path, anno3_instances, anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos):
    if os.path.exists(save_path):
        print(f'remove and update the {save_path}')
        os.remove(save_path)
    dialogs2otherinfo = collections.OrderedDict()
    for instance in anno3_instances:
        UtteranceId = instance[0].value
        if UtteranceId is None:
            continue
        dialog_id = '_'.join(UtteranceId.split('_')[:2])
        start_time, end_time, text, spk = instance[1:5]
        cl_text = clean_text(text.value)
        if dialogs2otherinfo.get(dialog_id) is None:
            dialogs2otherinfo[dialog_id] = [[UtteranceId, start_time.value, end_time.value, cl_text, spk.value.replace(' ', '')]]
        else:
            dialogs2otherinfo[dialog_id] += [[UtteranceId, start_time.value, end_time.value, cl_text, spk.value.replace(' ', '')]]
    # print(list(dialogs2otherinfo.keys()))
    new_instances = []
    new_instances.append(['UtteranceId', 'StartTime', 'EndTime', 'Text', 'Speaker', 'EmoAnnotator1', 'EmoAnnotator2', 'EmoAnnotator3', 'final_mul_emo', 'final_main_emo', 
                                            'CKappa(P1P2)', 'CKappa(P1P3)', 'CKappa(P2P3)', 'CKappa(P1Main)', 'CKappa(P2Main)', 'CKappa(P3Main)', 'NotSure',
                                            'Sim(P1P2)', 'Sim(P1P3)', 'Sim(P2P3)', 'Sim(P1Main)', 'Sim(P2Main)', 'Sim(P3Main)', 
                                            'Feedback', 'Sim(P1Feed)', 'Sim(P2Feed)', 'Sim(P3Feed)', 'P1Quality', 'P2Quality', 'P3Quality'])
    for dialog_id in dialogs2otherinfo.keys():
        otherinfos = dialogs2otherinfo[dialog_id]
        anno1_emos = anno1_dialogs2emos[dialog_id]
        anno2_emos = anno2_dialogs2emos[dialog_id]
        anno3_emos = anno3_dialogs2emos[dialog_id]
        assert len(otherinfos) == len(anno1_emos) == len(anno2_emos) == len(anno3_emos)
        for i in range(len(otherinfos)):
            temp = otherinfos[i]
            temp.append(','.join(anno1_emos[i]))
            temp.append(','.join(anno2_emos[i]))
            temp.append(','.join(anno3_emos[i]))
            new_instances.append(temp)
    write_xls(save_path, sheetname='sheet1', instances=new_instances)

def decision_allin_pool(anno1_emos, anno2_emos, anno3_emos):
    emos_pool = anno1_emos + anno2_emos + anno3_emos
    # print(f'emos pooling {emos_pool}')
    emo2count = Counter(emos_pool)
    count2emos = {}
    for emo in emo2count.keys():
        count = emo2count[emo]
        if count2emos.get(count) is None:
            count2emos[count] = [emo]
        else:
            count2emos[count] += [emo]
    # print(emo2count, count2emos)
    all_values = emo2count.values()
    if len(all_values) == sum(all_values):
        print(f'Warning all emotion are uniqe {emo2count}')
        mul_emos, main_emo = ['NotSure'], 'NotSure'
    else:
        # 一个情感最多出现三次
        count2_emos, count3_emos = [], []
        if count2emos.get(2) is not None:
            count2_emos.extend(count2emos[2])
        if count2emos.get(3) is not None:
            count3_emos.extend(count2emos[3])
        mul_emos, main_emo = [], None
        if len(count3_emos) == 1:
            main_emo =  count3_emos[0]
            mul_emos.append(main_emo)
            if len(count2_emos) > 0:
                mul_emos.extend(count2_emos)
                # print(f'{main_emo} appear 3 times and {count2_emos} appear 2 times')
            # else:
                # print(f'{main_emo} appear 3 times and other emos appear 0/1 times')
        elif len(count3_emos) > 1:
            # 随便出现次数多的第一情感
            mul_emos.extend(count3_emos)
            first_emos =  [anno1_emos[0], anno2_emos[0], anno3_emos[0]]
            first_emo2count = Counter(first_emos)
            for emo in first_emo2count:
                if first_emo2count[emo] > 1:
                    main_emo = emo
            if main_emo is None:
                main_emo = 'NotSure'
                print('Warning please check one, two or more emotion 3times but no main emo')
                # print(first_emo2count, count2emos)
            # print(f'{count3_emos} appear 3 times')
        else:
            # 没有出现三次的情感, 并且出现两次的只有一个情感, 那么该情感即为主情感也是多情感
            if len(count2_emos) == 1:
                main_emo = count2_emos[0]
                mul_emos = count2_emos
            else:
                # 出现次数都是0的已经讨论过了
                assert len(count2_emos) > 0
                mul_emos = count2_emos
                first_emos =  [anno1_emos[0], anno2_emos[0], anno3_emos[0]]
                first_emo2count = Counter(first_emos)
                for emo in first_emo2count:
                    if first_emo2count[emo] > 1:
                        main_emo = emo
                if main_emo is None:
                    main_emo = 'NotSure'
                    print('Warning please check one,  two or more emotion 2 times but no main emo')
                    # print(first_emo2count, count2emos)
                # print(f'{count2_emos} appear 2 times')
    return mul_emos, main_emo

def compute_fleiss_kappa(p1_emos, p2_emos, p3_emos):
    #  https://github.com/Shamya/FleissKappa
    # 按照格式进行整理，成 rate 矩阵 比如共100个样本，共5个标注类别，共有3个标注者（每一行的和是3），那么 rate[0][4] 表示这个样本中14个人都选择了类别5.
    from fleiss import fleissKappa
    emo_list = ['Happy', 'Neutral', 'Sad', 'Anger', 'Disgust', 'Fear', 'Surprise', 'Other']
    rate = []
    for i in range(len(p1_emos)):
        subject = [0] * len(emo_list)
        subject[emo_list.index(p1_emos[i])] += 1
        subject[emo_list.index(p2_emos[i])] += 1
        subject[emo_list.index(p3_emos[i])] += 1
        rate.append(subject)
    kappa = fleissKappa(rate, 3)
    return kappa

def compute_cohen_kappa_between_annotators(p1_emos, p2_emos, p3_emos, main_emos):
    # compute P1P2 P1P3 P2P3 P1Main P2Main P3Main
    emo_list = ['Happy', 'Neutral', 'Sad', 'Anger', 'Disgust', 'Fear', 'Surprise', 'Other', 'NotSure']
    p1_emos_labels, p2_emos_labels, p3_emos_labels, main_emos_labels = [], [], [], []
    for i in range(len(p1_emos)):
        p1_emos_labels.append(emo_list.index(p1_emos[i]))
        p2_emos_labels.append(emo_list.index(p2_emos[i]))
        p3_emos_labels.append(emo_list.index(p3_emos[i]))
        main_emos_labels.append(emo_list.index(main_emos[i]))
    assert len(p1_emos_labels) == len(p2_emos_labels) == len(p3_emos_labels) == len(main_emos_labels)
    p1p2 = sklearn.metrics.cohen_kappa_score(p1_emos_labels, p2_emos_labels)
    p1p3 = sklearn.metrics.cohen_kappa_score(p1_emos_labels, p3_emos_labels)
    p2p3 = sklearn.metrics.cohen_kappa_score(p2_emos_labels, p3_emos_labels)
    p1main = sklearn.metrics.cohen_kappa_score(p1_emos_labels, main_emos_labels)
    p2main = sklearn.metrics.cohen_kappa_score(p2_emos_labels, main_emos_labels)
    p3main = sklearn.metrics.cohen_kappa_score(p3_emos_labels, main_emos_labels)
    return p1p2, p1p3, p2p3, p1main, p2main, p3main

def get_sim_score(anno1_emo_labels, anno2_emo_labels):
    # 这里是原始的多情感标注
    assert len(anno1_emo_labels) == len(anno2_emo_labels)
    sim_count = 0
    for i in range(len(anno1_emo_labels)):
        anno1_es, anno2_es = anno1_emo_labels[i], anno2_emo_labels[i]
        if len(set(anno1_es).intersection(anno2_es)) > 0:
            sim_count += 1
    return round(sim_count/len(anno1_emo_labels), 2)

def get_sim_score_3p(anno1_emo_labels, anno2_emo_labels, anno3_emo_labels):
    # 这里是原始的多情感标注
    assert len(anno1_emo_labels) == len(anno2_emo_labels) == len(anno3_emo_labels)
    sim_count = 0
    for i in range(len(anno1_emo_labels)):
        anno1_es, anno2_es, anno3_es = anno1_emo_labels[i], anno2_emo_labels[i], anno3_emo_labels[i]
        if len(set(anno1_es).intersection(set(anno2_es), set(anno3_es))) > 0:
            sim_count += 1
    return round(sim_count/len(anno1_emo_labels), 2)

def compute_sim_between_annotators(p1_emos, p2_emos, p3_emos, main_emos,  dialog2num_utts=None):
    # p1_emos 每个元素都是多情感 list
    # cohen_kappa 计算不太直观，为了更好的对比差异性, 考虑到第二情感，很多时候并不能很好的把握二者的顺序，如果有匹配的即可
    # compute P1P2 P1P3 P2P3 P1Main P2Main P3Main P1Feedback P2Feedback P3Feedback
    emo_list = ['Happy', 'Neutral', 'Sad', 'Anger', 'Disgust', 'Fear', 'Surprise', 'Other', 'NotSure']
    p1_emos_labels, p2_emos_labels, p3_emos_labels, main_emos_labels = [], [], [], []
    for i in range(len(p1_emos)):
        p1_emos_labels.append([emo_list.index(temp) for temp in p1_emos[i]])
        p2_emos_labels.append([emo_list.index(temp) for temp in p2_emos[i]])
        p3_emos_labels.append([emo_list.index(temp) for temp in p3_emos[i]])
        main_emos_labels.append([emo_list.index(main_emos[i])])
    assert len(p1_emos_labels) == len(p2_emos_labels) == len(p3_emos_labels) == len(main_emos_labels)
    p1p2 = get_sim_score(p1_emos_labels, p2_emos_labels)
    p1p3 = get_sim_score(p1_emos_labels, p3_emos_labels)
    p2p3 = get_sim_score(p2_emos_labels, p3_emos_labels)
    p1main = get_sim_score(p1_emos_labels, main_emos_labels)
    p2main = get_sim_score(p2_emos_labels, main_emos_labels)
    p3main = get_sim_score(p3_emos_labels, main_emos_labels)

    if dialog2num_utts is not None:
        # dialog2num_utts provide num_utt of each dialog
        dialog2sims = collections.OrderedDict()
        current_num = 0
        for dialog_id in dialog2num_utts.keys():
            start = current_num
            end = current_num + dialog2num_utts[dialog_id]
            p1_emo_labels = p1_emos_labels[start:end]
            p2_emo_labels = p2_emos_labels[start:end]
            p3_emo_labels = p3_emos_labels[start:end]
            dialog_sim = get_sim_score_3p(p1_emo_labels, p2_emo_labels, p3_emo_labels)
            dialog2sims[dialog_id] = dialog_sim
            print(dialog_id, start, end, dialog_sim)
            current_num = end
    else:
        dialog2sims = None

    return p1p2, p1p3, p2p3, p1main, p2main, p3main, dialog2sims


def compute_sim_with_feedbacks(p1_emos, p2_emos, p3_emos, feedback_emos):
    # p1_emos 每个元素都是多情感 list
    # if no feedback then the emo is None, ignore all this
    emo_list = ['Happy', 'Neutral', 'Sad', 'Anger', 'Disgust', 'Fear', 'Surprise', 'Other']
    p1_emos_labels, p2_emos_labels, p3_emos_labels, feedback_emos_labels = [], [], [], []
    for i in range(len(p1_emos)):
        # 只跟Feedback的数据进行比较，作为Ground-truth
        if feedback_emos[i] is not None:
            p1_emos_labels.append([emo_list.index(temp) for temp in p1_emos[i]])
            p2_emos_labels.append([emo_list.index(temp) for temp in p2_emos[i]])
            p3_emos_labels.append([emo_list.index(temp) for temp in p3_emos[i]])
            feedback_emos_labels.append([emo_list.index(temp) for temp in feedback_emos[i]])
    assert len(p1_emos_labels) == len(p2_emos_labels) == len(p3_emos_labels) == len(feedback_emos_labels)
    p1feed = get_sim_score(p1_emos_labels, feedback_emos_labels)
    p2feed = get_sim_score(p2_emos_labels, feedback_emos_labels)
    p3feed = get_sim_score(p3_emos_labels, feedback_emos_labels)
    return p1feed, p2feed, p3feed


def get_final_decision(filepath, strategy_name='allin_pool', feedback_dialogs2emos=None, quality_scores=None):
    workbook = openpyxl.load_workbook(filepath)
    booksheet = workbook['sheet1']
    rows = booksheet.rows
    rows = [r for r in rows][1:]
    all_anno1_emos, all_anno2_emos, all_anno3_emos = [], [], []
    all_main_emos, all_p1_emos, all_p2_emos, all_p3_emos = [], [], [], []
    not_sure_count = 0
    all_feedback_emos = []
    exist_feedback_dialogs = []
    dialog2num_utts = collections.OrderedDict()
    for i in range(len(rows)):
        # print(f'[Debug] in utterance {rows[i][0].value}')
        anno1_emo_str, anno2_emo_str, anno3_emo_str = rows[i][5].value, rows[i][6].value, rows[i][7].value
        anno1_emos, anno2_emos, anno3_emos = anno1_emo_str.split(','), anno2_emo_str.split(','), anno3_emo_str.split(',')
        # print(anno1_emo_str, anno2_emo_str, anno3_emo_str)
        # print(anno1_emos, anno2_emos, anno3_emos)
        if strategy_name == 'allin_pool':
            mul_emos, main_emo = decision_allin_pool(anno1_emos, anno2_emos, anno3_emos)
        else:
            print(f'Error of strategy_name {strategy_name}')
        # start from 1 and skip 1
        update_row_idx = i+2
        booksheet.cell(row=update_row_idx, column=9).value= ','.join(mul_emos)
        booksheet.cell(row=update_row_idx, column=10).value= main_emo
        if main_emo == 'NotSure':
            not_sure_count += 1

        # 保存单标签
        all_main_emos.append(main_emo)
        all_p1_emos.append(anno1_emos[0])
        all_p2_emos.append(anno2_emos[0])
        all_p3_emos.append(anno3_emos[0])
        # 保存多标签
        all_anno1_emos.append(anno1_emos)
        all_anno2_emos.append(anno2_emos)
        all_anno3_emos.append(anno3_emos)
        
        if feedback_dialogs2emos is not None:
            utterance_id = rows[i][0].value
            dialog_id = '_'.join(utterance_id.split('_')[:2])
            if feedback_dialogs2emos.get(dialog_id) is not None:
                if dialog_id not in exist_feedback_dialogs:
                    exist_feedback_dialogs.append(dialog_id)
                    all_feedback_emos.extend(feedback_dialogs2emos[dialog_id])
            else:
                all_feedback_emos.append(None)

            if dialog2num_utts.get(dialog_id) is None:
                dialog2num_utts[dialog_id] = 1
            else:
                dialog2num_utts[dialog_id] += 1
    # compute the kappa between annotators and save
    p1p2, p1p3, p2p3, p1main, p2main, p3main = compute_cohen_kappa_between_annotators(all_p1_emos, all_p2_emos, all_p3_emos, all_main_emos)
    booksheet.cell(row=2, column=11).value= p1p2
    booksheet.cell(row=2, column=12).value= p1p3
    booksheet.cell(row=2, column=13).value= p2p3
    booksheet.cell(row=2, column=14).value= p1main
    booksheet.cell(row=2, column=15).value= p2main
    booksheet.cell(row=2, column=16).value= p3main
    booksheet.cell(row=2, column=17).value= not_sure_count

    # compute the simlairy between annotators and save
    sim_p1p2, sim_p1p3, sim_p2p3, sim_p1main, sim_p2main, sim_p3main, dailog2sim3p = compute_sim_between_annotators(
                                                    all_anno1_emos, all_anno2_emos, all_anno3_emos, all_main_emos, dialog2num_utts)
    booksheet.cell(row=2, column=18).value= sim_p1p2
    booksheet.cell(row=2, column=19).value= sim_p1p3
    booksheet.cell(row=2, column=20).value= sim_p2p3
    booksheet.cell(row=2, column=21).value= sim_p1main
    booksheet.cell(row=2, column=22).value= sim_p2main
    booksheet.cell(row=2, column=23).value= sim_p3main

    if dailog2sim3p is not None:
        current_column = 32
        for dialog_id in dailog2sim3p.keys():            
            booksheet.cell(row=1, column=current_column, ).value = dialog_id
            booksheet.cell(row=2, column=current_column).value= dailog2sim3p[dialog_id]
            current_column += 1

    if feedback_dialogs2emos is not None:
        for rowidx in range(len(all_feedback_emos)):
            if all_feedback_emos[rowidx] is not None:
                booksheet.cell(row=rowidx+2, column=24).value = ','.join(all_feedback_emos[rowidx])
        assert len(all_feedback_emos) == len(all_anno1_emos)            
        sim_p1feed, sim_p2feed, sim_p3feed = compute_sim_with_feedbacks(all_anno1_emos, all_anno2_emos, all_anno3_emos, all_feedback_emos)
        booksheet.cell(row=2, column=25).value= sim_p1feed
        booksheet.cell(row=2, column=26).value= sim_p2feed
        booksheet.cell(row=2, column=27).value= sim_p3feed

    if quality_scores is not None:
        current_column = 28
        booksheet.cell(row=2, column=current_column).value= quality_scores[0]
        booksheet.cell(row=2, column=current_column+1).value= quality_scores[1]
        booksheet.cell(row=2, column=current_column+2).value= quality_scores[2]

    fleiss_kappa = compute_fleiss_kappa(all_p1_emos, all_p2_emos, all_p3_emos)
    booksheet.cell(row=1, column=31).value= 'FleissKappa'
    booksheet.cell(row=2, column=31).value= fleiss_kappa
    # save
    workbook.save(filepath)
    return fleiss_kappa, not_sure_count

def get_anno_quality(admin_filepath):
    movie2quality = {}
    movie2annotators = {}
    all_instances = read_xls(admin_filepath, sheetname='电视剧列表', skip_rows=1)
    for instance in all_instances:
        # first row is movie_name
        movie_name = instance[0].value
        anno1_quality = instance[6].value
        anno1_quality_score = Counter(anno1_quality)['★']
        anno2_quality = instance[12].value
        anno2_quality_score = Counter(anno2_quality)['★']
        anno3_quality = instance[17].value
        anno3_quality_score = Counter(anno3_quality)['★']
        movie2quality[movie_name] = [anno1_quality_score, anno2_quality_score, anno3_quality_score]
        anno1_name = instance[5].value
        anno2_name = instance[11].value
        anno3_name = instance[16].value
        movie2annotators[movie_name] = [anno1_name, anno2_name, anno3_name]
    return movie2quality, movie2annotators

def get_feedback_dialog(anno_instances):
    # input: instances of one annotation filepath
    # get dialog with Feedback column
    dialog2feedbacks = collections.OrderedDict()
    for instance_id in range(len(anno_instances)):
        instance = anno_instances[instance_id]
        UtteranceId = instance[0].value
        if UtteranceId is None:
            continue
        dialog_id = '_'.join(UtteranceId.split('_')[:2])
        feedback = instance[7].value
        if dialog2feedbacks.get(dialog_id) is None:
            dialog2feedbacks[dialog_id] = [feedback]
        else:
            dialog2feedbacks[dialog_id] += [feedback]
    feedback_dialogs = []
    for dialog_id in dialog2feedbacks.keys():
        feedback_flag = False
        feedbacks = dialog2feedbacks[dialog_id]
        for feedback in feedbacks:
            if feedback is not None and len(feedback) > 1:
                feedback_flag = True
        if feedback_flag:
            feedback_dialogs.append(dialog_id)
    return feedback_dialogs

def get_feedback_ground_emos(anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos, anno1_feedback_dialogs, 
                                    anno2_feedback_dialogs, anno3_feedback_dialogs):
    # 由于随着后来的标注越来熟练，所以以P3优先, P1有些数据Check的比较多比较详细, 最后 P2
    feedback_dialogs2emos = collections.OrderedDict()
    for dialog_id in anno3_feedback_dialogs:
        feedback_dialogs2emos[dialog_id] = anno3_dialogs2emos[dialog_id]
    for dialog_id in anno1_feedback_dialogs:
        if feedback_dialogs2emos.get(dialog_id) is None:
            feedback_dialogs2emos[dialog_id] = anno1_dialogs2emos[dialog_id]
    for dialog_id in anno2_feedback_dialogs:
        if feedback_dialogs2emos.get(dialog_id) is None:
            feedback_dialogs2emos[dialog_id] = anno2_dialogs2emos[dialog_id]
    return feedback_dialogs2emos

def collections_low_sim_dialogs(movie_name, meta_fileapth, annotators):
    # head = ['UtteranceId', 'StartTime', 'EndTime', 'Text', 'Speaker', 'EmoAnnotator1', 'EmoAnnotator2', 'EmoAnnotator3', 'final_mul_emo', 'final_main_emo']
    workbook = openpyxl.load_workbook(meta_fileapth)
    booksheet = workbook['sheet1']
    rows = booksheet.rows
    rows = [r for r in rows]
    fleissKappa = float(rows[1][30].value)
    if fleissKappa > 0.45:
        threshold = 0.3
    else:
        threshold = 0.5
    low_sim_dialogs2score = collections.OrderedDict()
    for name, sim in zip(rows[0], rows[1]):
        if movie_name in name.value:
            if float(sim.value) < threshold:
                low_sim_dialogs2score[name.value] = sim.value
    instances = []
    instances.append([movie_name, annotators[0], annotators[1], annotators[2] ])
    for instance in rows[1:]:
        UtteranceId = instance[0].value
        if UtteranceId is None:
            continue
        dialog_id = '_'.join(UtteranceId.split('_')[:2])
        if low_sim_dialogs2score.get(dialog_id) is not None:
            temp = [s.value for s in instance]
            instances.append(temp[:10])
    return instances, low_sim_dialogs2score

def collections_low_sim_personIndex_dialogs(movies_names, movie2annotators, low_sim_filepath, low_sim_simple_filepath, 
                                                low_sim_simple_personindex_filepath):
    all_instances = [['UtteranceId', 'StartTime', 'EndTime', 'Text', 'Speaker', 'EmoAnnotator1', 'EmoAnnotator2', 'EmoAnnotator3', 
                                                'final_mul_emo', 'final_main_emo', 'simp1p2p3', 'P1', 'P2', 'P3']]
    all_instances_simple = [['movie_name', 'dialog_id', 'P1', 'P2', 'P3', 'sim_score']]
    for movie_name in movies_names:
        meta_fileapth = '/Users/jinming/Desktop/works/memoconv_final_labels/meta_{}.xlsx'.format(movie_name)
        annotators = movie2annotators[movie_name]
        instances, low_sim_dialogs2score = collections_low_sim_dialogs(movie_name, meta_fileapth, annotators)
        all_instances.extend(instances)
        for dialog_id in low_sim_dialogs2score:
            all_instances_simple.append([movie_name, dialog_id, annotators[0], annotators[1], annotators[2], low_sim_dialogs2score[dialog_id]])
        print(movie_name, len(low_sim_dialogs2score))
    write_xls(low_sim_filepath, 'sheet1', all_instances)
    write_xls(low_sim_simple_filepath, 'sheet1', all_instances_simple)
    # 按person进行排序～
    anno2idialogs = collections.OrderedDict()
    for instance in all_instances_simple:
        dialog_id, anno1, anno2, anno3 = instance[1].value, instance[2].value, instance[3].value, instance[4].value
        if anno2idialogs.get(anno1) is None:
            anno2idialogs[anno1] = [dialog_id]
        else:
            anno2idialogs[anno1] += [dialog_id]
        if anno2idialogs.get(anno2) is None:
            anno2idialogs[anno2] = [dialog_id]
        else:
            anno2idialogs[anno2] += [dialog_id]
        if anno2idialogs.get(anno3) is None:
            anno2idialogs[anno3] = [dialog_id]
        else:
            anno2idialogs[anno3] += [dialog_id]
    new_all_instances = []
    for annotator in anno2idialogs.keys():
        for dialog_id in anno2idialogs[annotator]:
            new_all_instances.append([annotator, dialog_id])
    write_xls(low_sim_simple_personindex_filepath, 'sheet1', new_all_instances)

def get_spk_anno_format(output_filepath, anno3_instances, movie_name):
    all_instances = []
    all_instances.append(['Actors'])
    all_instances.append(['Age(Child,Young,Mid,Old)'])
    all_instances.append(['Gender'])
    all_instances.append(['OtherName'])
    all_instances.append([])
    all_instances.append(['DialogId', 'StartTime', 'StartUtt', 'SpeakerA', 'SpeakerB'])
    dialog2exist = {}
    for instance in anno3_instances:
        UtteranceId = instance[0].value
        if UtteranceId is None:
            continue
        dialog_id = '_'.join(UtteranceId.split('_')[:2])
        feedback = instance[7].value
        if dialog2exist.get(dialog_id) is None:
            dialog2exist[dialog_id] = [feedback]
            all_instances.append([UtteranceId, instance[1].value, instance[3].value])
    wb = openpyxl.load_workbook(output_filepath)
    ws = wb.create_sheet(movie_name, 0)
    for each in all_instances:
        ws.append(each)
    wb.save(output_filepath)

def write_meta_json_info():
    '''
    {'dialogId':
        'spkA': xxx,
        'spkB': xxx, 
        'starttime_raw_episode': xxx
        'endtime_raw_episode': xxx
        {
            uttId:{
                'speaker': A
                'starttime': xxx,
                'endtime': xxx,
                'duration': xxx,
                'text': xxx,
                'final_mul_emos': [],
                'final_main_emo': [],
                'annotator1': [],
                'annotator2': [],
                'annotator3': [],
                'annotator4': []
            }
        }
    }
    '''
    pass


if __name__ == '__main__':
    if True:
        admin_file_path = '/Users/jinming/Desktop/works/memoconv_labels/Admin.xlsx'
        movie2quality, movie2annotators = get_anno_quality(admin_file_path)

    movies_names = read_file('movie_list.txt')
    movies_names = [movie_name.strip() for movie_name in movies_names]

    if False:
        for movie_name in movies_names:
            anno1_path = '/Users/jinming/Desktop/works/memoconv_labels/{}_anno1_done.xlsx'.format(movie_name)
            if not os.path.exists(anno1_path):
                continue
            anno2_path = '/Users/jinming/Desktop/works/memoconv_labels/{}_anno2_done.xlsx'.format(movie_name)
            anno3_path = '/Users/jinming/Desktop/works/memoconv_labels/{}_anno3_done.xlsx'.format(movie_name)

            print('Current movie {}'.format(movie_name))
            # 整理统计的结果
            result_filepath = '/Users/jinming/Desktop/works/memoconv_final_labels/statistic.xlsx'
            meta_fileapth = '/Users/jinming/Desktop/works/memoconv_final_labels/meta_{}.xlsx'.format(movie_name)

            # step1
            anno1_instances = read_xls(anno1_path, sheetname=movie_name, skip_rows=1)
            anno2_instances = read_xls(anno2_path, sheetname='工作表1', skip_rows=1)
            anno3_instances = read_xls(anno3_path, sheetname='工作表1', skip_rows=1)
            print('\t anno1 {} anno2 {} anno3 {}'.format(len(anno1_instances), len(anno2_instances), len(anno3_instances)))
            # step2 & 3
            print('\t checking dailog 1')
            anno1_num_dialogs, anno1_num_turns, anno1_num_utts, anno1_dialogid2num_turns, anno1_dialogs = check_dialog(anno1_instances)
            print('\t checking dailog 2')
            anno2_num_dialogs, anno2_num_turns, anno2_num_utts, anno2_dialogid2num_turns, anno2_dialogs = check_dialog(anno2_instances)
            print('\t checking dailog 3')
            anno3_num_dialogs, anno3_num_turns, anno3_num_utts, anno3_dialogid2num_turns, anno3_dialogs = check_dialog(anno3_instances)
            assert anno1_num_dialogs == len(anno1_dialogid2num_turns) == len(anno1_dialogs)
            print('\t dialog_num: anno1 {} anno2 {} anno3 {}'.format(anno1_num_dialogs, anno1_num_dialogs, anno1_num_dialogs))
            assert anno1_num_dialogs == anno2_num_dialogs == anno3_num_dialogs
            print('\t dialog_turns_num: anno1 {} anno2 {} anno3 {}'.format(anno1_num_turns, anno2_num_turns, anno3_num_turns))
            flag = (anno1_num_turns == anno2_num_turns == anno3_num_turns)
            if not flag:
                for dialog_id in anno1_dialogid2num_turns.keys():
                    sub_flag = (anno1_dialogid2num_turns[dialog_id] == anno2_dialogid2num_turns[dialog_id] == anno3_dialogid2num_turns[dialog_id])
                    if not sub_flag:
                        print('\t \t Error in dailog {}: anno1 {} anno2 {} anno3 {}'.format(dialog_id, anno1_dialogid2num_turns[dialog_id], 
                                        anno2_dialogid2num_turns[dialog_id], anno3_dialogid2num_turns[dialog_id]))
            print('\t dialog_utts_num: anno1 {} anno2 {} anno3 {}'.format(anno1_num_utts, anno2_num_utts, anno3_num_utts))
            assert anno1_num_utts == anno2_num_utts == anno3_num_utts
            # step4
            print('\t analyse emotion of dailog 1')
            anno1_total_emo_anno, anno1_dialogs2emos = analyse_dialog_emotions(anno1_instances)
            print('\t analyse emotion of dailog 2')
            anno2_total_emo_anno, anno2_dialogs2emos = analyse_dialog_emotions(anno2_instances)
            print('\t analyse emotion of dailog 3')
            anno3_total_emo_anno, anno3_dialogs2emos = analyse_dialog_emotions(anno3_instances)
            print('\t total emo annos: anno1 {} anno2 {} anno3 {}'.format(anno1_total_emo_anno, anno2_total_emo_anno, anno3_total_emo_anno))
            p3_num_total_same, p3_num_main_same, p3_num_part_same = analyse_emotion_three_consistency(anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos)
            print('\t three people emo consistency: total_same {} main_same {} part_same {}'.format(p3_num_total_same, p3_num_main_same, p3_num_part_same))
            p2_num_total_same, p2_num_main_same, p2_num_part_same = analyse_emotion_two_consistency(anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos)
            print('\t two people emo consistency: total_same {} main_same {} part_same {}'.format(p2_num_total_same, p2_num_main_same, p2_num_part_same))

            if False:
                # step5, 以anno3为标注，整理三个人的情感标注信息
                get_clean_text_dialog(meta_fileapth, anno3_instances, anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos)

            if False:
                anno1_feedback_dialogs = get_feedback_dialog(anno1_instances)
                anno2_feedback_dialogs = get_feedback_dialog(anno2_instances)
                anno3_feedback_dialogs = get_feedback_dialog(anno3_instances)
                print('feedback_dialogs anno1 {} anno2 {} anno3 {}'.format(anno1_feedback_dialogs, anno2_feedback_dialogs, anno3_feedback_dialogs))
                feedback_dialogs2emos = get_feedback_ground_emos(anno1_dialogs2emos, anno2_dialogs2emos, anno3_dialogs2emos, anno1_feedback_dialogs, anno2_feedback_dialogs, anno3_feedback_dialogs)

            if False:
                # step6, 根据step6整理的结果，尝试不同的决策策略
                fleiss_kappa, not_sure_count = get_final_decision(meta_fileapth, strategy_name='allin_pool', feedback_dialogs2emos=feedback_dialogs2emos, quality_scores=movie2quality[movie_name])

            if False:
                # step7: get the statistic of the movie 
                instance = get_movie_statistic(movie_name, anno3_num_dialogs, anno3_num_turns, anno3_num_utts, 
                                            p3_num_total_same, p3_num_main_same, p3_num_part_same,
                                            p2_num_total_same, p2_num_main_same, p2_num_part_same, 
                                            fleiss_kappa, not_sure_count)
                write_xls_oneline(result_filepath, instance)

    if False:
        low_sim_filepath = '/Users/jinming/Desktop/works/memoconv_final_labels/low_sim_statistic.xlsx'
        low_sim_simple_filepath = '/Users/jinming/Desktop/works/memoconv_final_labels/low_sim_statistic_simple.xlsx'
        low_sim_simple_personindex_filepath = '/Users/jinming/Desktop/works/memoconv_final_labels/low_sim_statistic_PIndex_simple.xlsx'
        collections_low_sim_personIndex_dialogs(movies_names, movie2annotators, low_sim_filepath, low_sim_simple_filepath, low_sim_simple_personindex_filepath)
    
    if False:
        spk_format_filepath = '/Users/jinming/Desktop/works/memoconv_final_labels/dialogSpkAnno.xlsx'
        for movie_name in movies_names:
            anno3_path = '/Users/jinming/Desktop/works/memoconv_labels/{}_anno3_done.xlsx'.format(movie_name)
            anno3_instances = read_xls(anno3_path, sheetname='工作表1', skip_rows=1)
            get_spk_anno_format(spk_format_filepath, anno3_instances, movie_name)